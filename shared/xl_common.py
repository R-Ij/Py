#!/usr/bin/env python
# -*- coding: utf-8 -*-
import datetime
import openpyxl as pxl
from openpyxl.drawing.image import Image as Dr
from contextlib import contextmanager
import shutil
import os
import xlrd
from shared import w_log as wl
from shared import common as s_cmn
# from shared import constant as const


class XLCommon:

    def __init__(self):
        self.cls_name = 'XLCommon-'

        # Excelテンプレートパス(デフォルト)
        self.file_pass = ''
        self.file_name = ''
        self.tmp_pass = ''
        self.save_pass = ''
        self.dt_now = datetime.datetime.now().strftime('%Y%m%d')

        # ブックの値(デフォルト)
        self.workbook = None
        self.worksheet = ''
        self.work_cell = 'A1'

        # 画像のサイズ
        self.img_size = 280

        self.cmn = s_cmn.Common()
        # self.const = const.Constant()

    @contextmanager
    def oxl_b(self, **kwargs):
        """
        openpyxlでExcelブックをコピーして開く関数(クローズも含む)
        :yield: workbook
        """
        log_name = self.cls_name + 'oxl_b'
        full_pass = self.cmn.dir_end_check(self.file_pass) + self.file_name
        tmp_full_pass = self.cmn.dir_end_check(self.tmp_pass) + self.file_name
        save_full_pass = self.cmn.dir_end_check(self.save_pass) + '【' + self.dt_now + '】' + self.file_name
        try:
            shutil.copy(full_pass, self.tmp_pass)
            self.workbook = pxl.load_workbook(tmp_full_pass)
            yield self.workbook

            # ブックを保存
            if self.workbook is not None:
                self.workbook.save(save_full_pass)

        except Exception as e:
            wl.Log().err_log(e, log_name)
            return False

        finally:
            # クローズ処理
            if self.workbook is not None:
                self.workbook.close()
            if kwargs['del_flg']:
                os.remove(tmp_full_pass)

    def xlr_data_get(self, **kwargs):
        """
        xlrdでセルの値を取得する関数(数式の値も取得する)
        :param kwargs: row: 行
                        col: 列
        :yield: セルの値
        """
        log_name = self.cls_name + 'xlr_data_get'
        try:
            full_pass = self.cmn.dir_end_check(self.file_pass) + self.file_name
            with xlrd.open_workbook(full_pass) as wb:
                if self.worksheet != '':
                    ws = wb.sheet_by_name(self.worksheet)
                else:
                    ws = wb.sheet_by_index(0)

                # 0スタートのため、-1の補正をかける
                wc = ws.cell(kwargs['row'] - 1, kwargs['col'] - 1)
            return wc.value

        except Exception as e:
            wl.Log().err_log(e, log_name)
            return ''

    def xl_print(self, file_pass):
        """
        Excelファイルに画像ファイルを貼り付ける関数
        :param   file_pass:  画像ファイルのパス(ファイル名込み)
        :return: boolean
        """
        log_name = self.cls_name + 'xl_print'

        try:
            wb = self.workbook

            if self.worksheet != '':
                ws = wb[self.worksheet]
            else:
                ws = wb.worksheets[0]

            wc = self.work_cell
            img = Dr(file_pass)
            img.width = img.height = self.img_size
            img.anchor = wc
            ws.add_image(img)
            return True

        except Exception as e:
            wl.Log().err_log(e, log_name)
            return False

    def xl_data_get(self):
        """
        openpyxlでセルの値を取得する関数(数式の値は取得できない)
        :return: セルの値
        """
        log_name = self.cls_name + 'xl_data_get'

        try:
            wb = self.workbook
            ws = wb[self.worksheet]
            wc = ws[self.work_cell]
            return wc.value

        except Exception as e:
            wl.Log().err_log(e, log_name)
            return ''

    def sheet_n(self):
        """
        openpyxlでセルの値を取得する関数(数式の値は取得できない)
        :return: セルの値
        """
        log_name = self.cls_name + 'sheet_n'

        try:
            wb = self.workbook
            ws = wb[self.worksheet]
            wc = ws[self.work_cell]
            return wc.value

        except Exception as e:
            wl.Log().err_log(e, log_name)
            return ''


if __name__ == '__main__':
    xl = XLCommon()
    print(xl.xlr_data_get(row=3, col=6))
