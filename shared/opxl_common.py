#!/usr/bin/env python
# -*- coding: utf-8 -*-
import datetime
import openpyxl as pxl
from openpyxl.drawing.image import Image as Dr
from contextlib import contextmanager
import shutil
import os
from shared import w_log as wl
from shared import common as s_cmn
import shared.constant as const
import shared.err as err


class XLCommon:

    def __init__(self):
        self.cls_name = 'oXL-'

        # Excelテンプレートパス(デフォルト)
        self.file_path = ''
        self.file_name = ''
        self.tmp_path = ''
        self.save_path = ''
        self.dt_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S%f')

        # ブックの値(デフォルト)
        self.workbook = None
        self.worksheet = None
        self.work_cell = 'A1'

        # 画像のサイズ
        self.img_size = 280

        self.cmn = s_cmn.Common()
        self.const = const.Constant()
        self.err = err.Err()

    @contextmanager
    def oxl_b(self, del_flg=True):
        """
        OpenPyXLでExcelブックをコピーして開く関数(クローズも含む)
        :param del_flg: 終了後にコピーファイルを削除するかどうか
        :yield: workbook
        """
        log_name = self.cls_name + 'oxl_b'
        full_path = self.cmn.dir_end_check(self.file_path) + self.file_name
        tmp_full_path = self.cmn.dir_end_check2(self.tmp_path) + self.file_name
        save_full_path = self.cmn.dir_end_check(self.save_path) + '【' + self.dt_now + '】' + self.file_name
        try:
            shutil.copy(full_path, self.tmp_path)
            self.workbook = pxl.load_workbook(tmp_full_path)
            yield self.workbook

            # ブックを保存
            if self.workbook is not None:
                self.workbook.save(save_full_path)

        except Exception as e:
            wl.Log().o_log(self.const.ERR_MESSAGE + str(e), log_name)
            self.err.print_ex(log_name)
            return False

        finally:
            # クローズ処理
            if self.workbook is not None:
                self.workbook.close()
            if del_flg:
                os.remove(tmp_full_path)

    def xl_print(self, file_path):
        """
        Excelファイルに画像ファイルを貼り付ける関数
        :param   file_path:  画像ファイルのパス(ファイル名込み)
        :return: boolean
        """
        log_name = self.cls_name + 'xl_print'

        try:
            wb = self.workbook

            if self.worksheet != '':
                ws = wb[self.worksheet]
            else:
                ws = wb.worksheets[0]

            wc = self.work_cell
            img = Dr(file_path)
            img.width = img.height = self.img_size
            img.anchor = wc
            ws.add_image(img)
            return True

        except Exception as e:
            wl.Log().o_log(self.const.ERR_MESSAGE + str(e), log_name)
            self.err.print_ex(log_name)
            return False

    def xl_data_get(self):
        """
        OpenPyXLでセルの値を取得する関数(数式の値は取得できない)
        :return: セルの値
        """
        log_name = self.cls_name + 'xl_data_get'

        try:
            wb = self.workbook
            ws = wb[self.worksheet]
            wc = ws[self.work_cell]
            return wc.value

        except Exception as e:
            wl.Log().o_log(self.const.ERR_MESSAGE + str(e), log_name)
            self.err.print_ex(log_name)
            return ''

    def sheet_n(self):
        """
        OpenPyXLでセルの値を取得する関数(数式の値は取得できない)
        :return: セルの値
        """
        log_name = self.cls_name + 'sheet_n'

        try:
            wb = self.workbook
            ws = wb[self.worksheet]
            wc = ws[self.work_cell]
            return wc.value

        except Exception as e:
            wl.Log().o_log(self.const.ERR_MESSAGE + str(e), log_name)
            self.err.print_ex(log_name)
            return ''

    def oxl_get_sheet_name(self):
        """
        OpenPyXLでシート名を一覧で取得する関数
        """
        if self.workbook is not None:
            for name in self.workbook.get_sheet_names():
                yield name
        else:
            return None


if __name__ == '__main__':
    xl = XLCommon()
    # print(xl.xlr_data_get(row=3, col=6))

